import openpyxl

excel = openpyxl.load_workbook("pop_area_2009_2.xlsx")
sheet_name = excel.sheetnames[0]

sheet = excel[sheet_name]
for i in range(5,10):
    print(sheet.cell(row=i, column=4).value)